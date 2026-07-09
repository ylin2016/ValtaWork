"""Export a per-listing end-balance workbook for a period.

Reproduces the `output/<period>/end_balances_<period>.xlsx` report from the DB's
latest statement run for that period. Columns mirror the original May workbook:

  Listing | Owner(s) | VRP Net Income | My End Balance | My Owner Net Revenue | My Expense
          | Difference (mine - VRP) | Match status | Matched VRP listing

Where (from statement_property_totals):
  My End Balance       = amount_due_to_owner
  My Expense           = total_expenses
  My Owner Net Revenue = amount_due_to_owner - total_expenses

VRP columns are populated only if a production owner-statements CSV is passed via
--vrp (e.g. data/owner-statements_<period>_VRP.csv); otherwise they are left blank
and every row is flagged "MINE ONLY (no VRP file)".

Usage:
  python -m src.export_end_balances --period 2026-06 [--vrp data/owner-statements_2026-06_VRP.csv]
"""
import argparse
import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from src.config import load_config
from src.db import connect
from src.listing_filter import _alias


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _money(s):
    """Parse a VRP money cell: '$6,089.76 ', '(draft) $6,089.76', '' -> float."""
    s = re.sub(r"\(draft\)", "", s or "", flags=re.I)
    s = s.replace("$", "").replace(",", "").strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")  # accounting negatives
    s = s.strip("()")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def load_vrp(path):
    """Return {lookup_key: (raw_listing, net_income)} from a VRP CSV.

    Each listing is keyed both by its normalized name and by the alias form
    (listing_filter._alias), so VRP labels that differ from my property naming
    (e.g. "Cottage 8" -> osbr8, "Cottage All (OSBR)" -> osbr) still match.
    """
    vrp = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            listing = (row.get("Listing") or "").strip()
            if not listing:
                continue
            entry = (listing, _money(row.get("Net Income")))
            n = _norm(listing)
            vrp[n] = entry
            vrp.setdefault(_alias(n), entry)  # don't clobber a direct-name key
    return vrp


def latest_run_id(conn, period_start: str):
    row = conn.execute(
        "SELECT run_id FROM statement_runs WHERE period_start=? "
        "ORDER BY created_at DESC LIMIT 1",
        (period_start,),
    ).fetchone()
    return row["run_id"] if row else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--period", required=True, help="YYYY-MM")
    ap.add_argument("--config", default="config.yml")
    ap.add_argument("--vrp", default=None, help="Optional VRP owner-statements CSV to compare against")
    ap.add_argument("--output-dir", default="output")
    args = ap.parse_args()

    cfg = load_config(args.config)
    conn = connect(cfg["app"]["db_path"])

    period = args.period
    period_start = f"{period}-01"
    run_id = latest_run_id(conn, period_start)
    if not run_id:
        raise SystemExit(f"No statement run found for {period}")

    # VRP file: use --vrp if given, else auto-detect the conventional path.
    vrp_path = args.vrp
    if vrp_path is None:
        default_vrp = Path("data") / f"owner-statements_{period}_VRP.csv"
        if default_vrp.exists():
            vrp_path = str(default_vrp)
    vrp = load_vrp(vrp_path) if vrp_path else {}
    matched_vrp = set()
    if vrp_path:
        n_vrp = len({_norm(raw) for raw, _ in vrp.values()})
        print(f"Comparing against VRP summary: {vrp_path} ({n_vrp} listings)")

    rows = conn.execute(
        """SELECT spt.property_id, p.property_name, o.owner_name,
                  spt.amount_due_to_owner AS end_bal,
                  spt.total_expenses AS expense
             FROM statement_property_totals spt
             JOIN properties p ON spt.property_id = p.property_id
             JOIN owners o ON p.owner_id = o.owner_id
            WHERE spt.run_id = ?
            ORDER BY p.property_name""",
        (run_id,),
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"All Listings {period}"

    ws.append([f"Owner Statement End-Balance Comparison — {period}"])
    ws.append([
        "Mine = latest run from statement_property_totals. VRP = production owner-statements CSV. "
        "Difference = mine − VRP."
    ])
    ws.append([])
    header = [
        "Listing", "Owner(s)", "VRP Net Income", "My End Balance",
        "My Owner Net Revenue", "My Expense", "Difference (mine − VRP)",
        "Match status", "Matched VRP listing",
    ]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for r in rows:
        end_bal = round(float(r["end_bal"] or 0), 4)
        expense = round(float(r["expense"] or 0), 4)
        owner_net_rev = round(end_bal - expense, 4)

        name_key = _norm(r["property_name"])
        id_key = _norm(r["property_id"])
        v = vrp.get(name_key) or vrp.get(id_key)
        if not vrp_path:
            vrp_ni, diff, status, matched = None, None, "MINE ONLY (no VRP file)", None
        elif v is None:
            vrp_ni, diff, status, matched = None, None, "MINE ONLY (not in VRP)", None
        else:
            matched_label, vrp_ni = v
            matched_vrp.add(_norm(matched_label))
            diff = round(end_bal - vrp_ni, 4) if vrp_ni is not None else None
            status = "MATCH" if diff is not None and abs(diff) < 0.01 else "DIFFER"
            matched = matched_label

        ws.append([
            r["property_name"], r["owner_name"], vrp_ni, end_bal,
            owner_net_rev, expense, diff, status, matched,
        ])

    # VRP listings with no counterpart on my side (dedup dual keys by raw name).
    seen = set()
    for raw_listing, vrp_ni in sorted(vrp.values()):
        n = _norm(raw_listing)
        if n in matched_vrp or n in seen:
            continue
        seen.add(n)
        ws.append([
            raw_listing, None, vrp_ni, None, None, None, None,
            "VRP ONLY (not in mine)", raw_listing,
        ])

    for col, width in {
        "A": 28, "B": 26, "C": 15, "D": 16, "E": 20, "F": 14, "G": 22, "H": 26, "I": 22,
    }.items():
        ws.column_dimensions[col].width = width

    out_dir = Path(args.output_dir) / period
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"end_balances_{period}.xlsx"
    wb.save(out)
    total = sum(float(r["end_bal"] or 0) for r in rows)
    print(f"Wrote {len(rows)} listings to {out} (total end balance ${total:,.2f})")


if __name__ == "__main__":
    main()
