from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────
_DARK_BLUE  = "1F4E79"
_MED_BLUE   = "2E74B5"
_LIGHT_BLUE = "BDD7EE"
_GREEN_FILL = "E2EFDA"
_ALT_FILL   = "F5F5F5"
_WHITE      = "FFFFFF"

_SECTION_FILL = PatternFill("solid", fgColor=_DARK_BLUE)
_COL_HDR_FILL = PatternFill("solid", fgColor=_MED_BLUE)
_SUMM_FILL    = PatternFill("solid", fgColor=_LIGHT_BLUE)
_TOTAL_FILL   = PatternFill("solid", fgColor=_GREEN_FILL)
_ALT          = PatternFill("solid", fgColor=_ALT_FILL)

_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(bottom=_THIN)

CURRENCY = '$#,##0.00'
NCOLS    = 11  # A–K
FONT_SIZE = 14

# ── helpers ───────────────────────────────────────────────────────────────────
def _cell(ws, row, col, value=None, bold=False, size=FONT_SIZE, color=None,
          bg=None, fmt=None, align="left", wrap=False, merge_to=None):
    c = ws.cell(row=row, column=col, value=value)
    kw = {"bold": bold, "size": size}
    if color:
        kw["color"] = color
    c.font = Font(**kw)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg) if isinstance(bg, str) else bg
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if merge_to and merge_to > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c

def _section_row(ws, row, title):
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, color=_WHITE, size=FONT_SIZE)
    c.fill = _SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    return row + 1

def _col_header_row(ws, row, headers):
    ws.row_dimensions[row].height = 16
    for i in range(1, NCOLS + 1):
        c = ws.cell(row=row, column=i)
        if i <= len(headers):
            c.value = headers[i - 1]
        c.font = Font(bold=True, color=_WHITE, size=FONT_SIZE)
        c.fill = _COL_HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1

def _total_row(ws, row, label, amounts_by_col):
    ws.row_dimensions[row].height = 16
    for col in range(1, NCOLS + 1):
        ws.cell(row=row, column=col).fill = _TOTAL_FILL
    _cell(ws, row, 1, label, bold=True, bg=_TOTAL_FILL)
    for col, val in amounts_by_col.items():
        _cell(ws, row, col, val, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
    return row + 1

def _spacer(ws, row, height=6):
    ws.row_dimensions[row].height = height
    return row + 1

def _fmt_dates(checkin_str, checkout_str):
    try:
        ci = datetime.strptime(str(checkin_str)[:10], "%Y-%m-%d")
    except Exception:
        return str(checkin_str or "")
    if checkout_str:
        try:
            co = datetime.strptime(str(checkout_str)[:10], "%Y-%m-%d")
            nights = (co - ci).days
            return f"{ci.strftime('%d. %b.')} - {co.strftime('%d. %b. %Y')} / {nights} nights"
        except Exception:
            pass
    return ci.strftime("%d. %b. %Y")

# ── legacy template creator (used by init-db) ────────────────────────────────
def create_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Setup"
    for i, (k, v) in enumerate([
        ("Statement Period Start", ""), ("Statement Period End", ""),
        ("Property ID", ""), ("Property Name", ""), ("Owner Name", ""),
        ("Owner Email", ""), ("QBO Class ID", ""), ("Statement Version", "v1"),
    ], start=1):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    wb.save(path)

# ── main statement writer ────────────────────────────────────────────────────
def write_statement(output_path: str, period: str, property_info: dict,
                    owner_info: dict, pm_fee_rate: float, bookings: list,
                    other_income: list, expenses_by_subcat: dict, totals: dict,
                    owner_pays_cleaning: bool = False, owner_pays_supplies: bool = False,
                    owner_pays_taxes: bool = False):
    """
    Generate a single-sheet owner statement matching the PDF format.

    bookings: list of dicts with keys:
      - booking_id, guest_name, checkin, checkout, net_revenue
      - booking_channel_fee, booking_stripe_fee (negative values)
      - owner_cleaning_cost, owner_tax_cost (negative values)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"

    # column widths  A     B     C     D     E     F     G     H     I     J     K
    for i, w in enumerate([18, 15, 28, 18, 18, 18, 18, 18, 18, 18, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    period_dt    = datetime.strptime(period, "%Y-%m")
    period_label = period_dt.strftime("%B %Y")
    starting_bal = float(totals.get("starting_balance", 0))
    net_income   = float(totals.get("net_income", 0))
    current_bal  = starting_bal + net_income

    row = 1

    # ── HEADER ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    _cell(ws, row, 1, "Valta Realty", bold=True, size=16)
    _cell(ws, row, 5, f"{period_label} - Summary",
          bold=True, color=_WHITE, size=14, bg=_DARK_BLUE, align="center", merge_to=11)
    row += 1

    _cell(ws, row, 1, "4027 Beach Drive Southwest, Seattle, WA 98116", size=FONT_SIZE, color="595959")
    _cell(ws, row, 5, "Starting Balance", size=FONT_SIZE)
    _cell(ws, row, 11, starting_bal, size=FONT_SIZE, fmt=CURRENCY, align="right")
    row += 1

    _cell(ws, row, 1, "contact@valtarealty.com", size=FONT_SIZE, color="595959")
    _cell(ws, row, 5, "Net Income", bold=True, size=FONT_SIZE)
    _cell(ws, row, 11, net_income, bold=True, size=FONT_SIZE, fmt=CURRENCY, align="right")
    row += 1

    _cell(ws, row, 5, "Current Balance", size=FONT_SIZE)
    _cell(ws, row, 11, current_bal, size=FONT_SIZE, fmt=CURRENCY, align="right")
    row += 1

    ws.row_dimensions[row].height = 18
    _cell(ws, row, 1, property_info.get("property_name", ""), bold=True, size=14, merge_to=4)
    _cell(ws, row, 5, "Owner Payout", size=FONT_SIZE)
    _cell(ws, row, 11, net_income, size=FONT_SIZE, fmt=CURRENCY, align="right")
    row += 1

    _cell(ws, row, 1, "Owner", bold=True, size=FONT_SIZE, color="595959")
    _cell(ws, row, 2, owner_info.get("owner_name", ""), size=FONT_SIZE)
    _cell(ws, row, 5, "Ending Balance", bold=True, size=FONT_SIZE)
    _cell(ws, row, 11, current_bal - net_income, bold=True, size=FONT_SIZE, fmt=CURRENCY, align="right")
    row += 1

    _cell(ws, row, 1, "Email", bold=True, size=FONT_SIZE, color="595959")
    _cell(ws, row, 2, owner_info.get("owner_email", ""), size=FONT_SIZE)
    row += 1

    row = _spacer(ws, row)

    # ── NET REVENUE SECTION ───────────────────────────────────────────────────
    if bookings:
        row = _section_row(ws, row, "Net Revenue Section")

        headers = [
            "Reservation Dates", "Confirmation Code", "Guest Name",
            "Net Rental Revenue", "Channel Fee", "Stripe Fee"
        ]
        if owner_pays_cleaning:
            headers.append("Owner Cleaning Fee")
        if owner_pays_taxes:
            headers.append("Tax Paid to Owner")
        headers.extend(["Management Commission", "Net Owner Revenue", "Commission %"])

        row = _col_header_row(ws, row, headers)

        total_net_rental = total_channel = total_stripe = 0.0
        total_cleaning = total_tax = total_comm = total_owner = 0.0

        for i, b in enumerate(bookings):
            ws.row_dimensions[row].height = 15
            bg = _ALT if i % 2 else None

            net_rental = float(b.get("net_revenue") or 0)
            channel_fee = float(b.get("booking_channel_fee") or 0)
            stripe_fee = float(b.get("booking_stripe_fee") or 0)
            cleaning_fee = float(b.get("owner_cleaning_cost") or 0) if owner_pays_cleaning else 0.0
            tax_paid = float(b.get("owner_tax_cost") or 0) if owner_pays_taxes else 0.0

            comm = -net_rental * pm_fee_rate
            owner_rev = net_rental + channel_fee + stripe_fee + cleaning_fee + tax_paid + comm

            col = 1
            _cell(ws, row, col, _fmt_dates(b["checkin"], b["checkout"]), bg=bg)
            col += 1
            _cell(ws, row, col, b["booking_id"] or "", bg=bg, align="center")
            col += 1
            _cell(ws, row, col, b["guest_name"] or "", bg=bg)
            col += 1
            _cell(ws, row, col, net_rental, fmt=CURRENCY, align="right", bg=bg)
            col += 1
            _cell(ws, row, col, channel_fee, fmt=CURRENCY, align="right", bg=bg)
            col += 1
            _cell(ws, row, col, stripe_fee, fmt=CURRENCY, align="right", bg=bg)
            col += 1
            if owner_pays_cleaning:
                _cell(ws, row, col, cleaning_fee, fmt=CURRENCY, align="right", bg=bg)
                col += 1
            if owner_pays_taxes:
                _cell(ws, row, col, tax_paid, fmt=CURRENCY, align="right", bg=bg)
                col += 1
            _cell(ws, row, col, comm, fmt=CURRENCY, align="right", bg=bg)
            col += 1
            _cell(ws, row, col, owner_rev, fmt=CURRENCY, align="right", bg=bg)
            col += 1
            _cell(ws, row, col, -pm_fee_rate, fmt="0%", align="center", bg=bg)

            total_net_rental += net_rental
            total_channel += channel_fee
            total_stripe += stripe_fee
            total_cleaning += cleaning_fee
            total_tax += tax_paid
            total_comm += comm
            total_owner += owner_rev
            row += 1

        # Total row
        ws.row_dimensions[row].height = 16
        for col in range(1, NCOLS + 1):
            ws.cell(row=row, column=col).fill = _TOTAL_FILL
        _cell(ws, row, 1, "Total", bold=True, bg=_TOTAL_FILL)
        col = 4
        _cell(ws, row, col, total_net_rental, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        col += 1
        _cell(ws, row, col, total_channel, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        col += 1
        _cell(ws, row, col, total_stripe, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        col += 1
        if owner_pays_cleaning:
            _cell(ws, row, col, total_cleaning, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
            col += 1
        if owner_pays_taxes:
            _cell(ws, row, col, total_tax, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
            col += 1
        _cell(ws, row, col, total_comm, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        col += 1
        _cell(ws, row, col, total_owner, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        col += 1
        _cell(ws, row, col, -pm_fee_rate, bold=True, fmt="0%", align="center", bg=_TOTAL_FILL)
        row += 1
        row = _spacer(ws, row)

    # ── OTHER INCOME ──────────────────────────────────────────────────────────
    if other_income:
        row = _section_row(ws, row, "Other Credits")
        row = _col_header_row(ws, row, ["Date", "Description", "Type", "Amount"])

        total_oi = 0.0
        for i, oi in enumerate(other_income):
            ws.row_dimensions[row].height = 15
            bg = _ALT if i % 2 else None
            amt = float(oi["amount"] or 0)
            _cell(ws, row, 1, oi["posting_date"], bg=bg)
            _cell(ws, row, 2, (oi["description"] or "")[:200], bg=bg, wrap=True)
            _cell(ws, row, 3, oi["subcategory"] or "", bg=bg)
            _cell(ws, row, 4, amt, fmt=CURRENCY, align="right", bg=bg)
            total_oi += amt
            row += 1

        row = _total_row(ws, row, "Total:", {4: total_oi})
        row = _spacer(ws, row)

    # ── EXPENSE SECTIONS ──────────────────────────────────────────────────────
    ORDER = ["Repairs & Maintenance", "Cleaning Labor", "Supplies", "Utilities",
             "Internet/Cable", "HOA", "Insurance", "Property Taxes", "Other Expense"]

    def _sort_key(s):
        try:
            return ORDER.index(s)
        except ValueError:
            return len(ORDER)

    for subcat in sorted(expenses_by_subcat.keys(), key=_sort_key):
        lines = expenses_by_subcat[subcat]
        if not lines:
            continue

        # Skip supplies for owner-paid properties
        if subcat == "Supplies" and owner_pays_supplies:
            continue

        row = _section_row(ws, row, subcat)

        # Expense header: Date | Description(B-C merged) | Type(D-E merged) | (F) | Amount(G)
        ws.row_dimensions[row].height = 16
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=row, column=col)
            c.fill = _COL_HDR_FILL
            c.font = Font(bold=True, color=_WHITE, size=FONT_SIZE)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).value = "Date"
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2).value = "Description"
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.cell(row=row, column=4).value = "Type"
        ws.cell(row=row, column=7).value = "Amount"
        row += 1

        total_exp = 0.0
        for i, exp in enumerate(lines):
            ws.row_dimensions[row].height = 15
            bg = _ALT if i % 2 else None
            amt = float(exp["amount"] or 0)
            vendor_type = exp["qbo_account"] or exp["vendor_customer"] or ""

            _cell(ws, row, 1, exp["posting_date"], bg=bg)

            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            c = ws.cell(row=row, column=2)
            c.value = (exp["description"] or "")[:200]
            c.font = Font(size=FONT_SIZE)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            if bg:
                c.fill = bg

            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            c = ws.cell(row=row, column=4)
            c.value = vendor_type
            c.font = Font(size=FONT_SIZE)
            c.alignment = Alignment(vertical="center")
            if bg:
                c.fill = bg

            if bg:
                ws.cell(row=row, column=6).fill = bg

            _cell(ws, row, 7, amt, fmt=CURRENCY, align="right", bg=bg)
            total_exp += amt
            row += 1

        # Total row
        ws.row_dimensions[row].height = 16
        for col in range(1, NCOLS + 1):
            ws.cell(row=row, column=col).fill = _TOTAL_FILL
        _cell(ws, row, 1, "Total:", bold=True, bg=_TOTAL_FILL)
        _cell(ws, row, 7, total_exp, bold=True, fmt=CURRENCY, align="right", bg=_TOTAL_FILL)
        row += 1
        row = _spacer(ws, row)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
